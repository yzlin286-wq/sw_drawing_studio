"""v2.3 Task 4: 视觉审计报告生成器

将审计结果导出为 Excel/JSON/文本报告。
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from app.services.visual_audit_service import AuditResult


class VisualAuditReporter:
    """视觉审计报告生成器"""

    def __init__(self, audit_results: list[AuditResult]):
        """初始化报告生成器

        Args:
            audit_results: 审计结果列表
        """
        self.audit_results = audit_results

    @staticmethod
    def _issue_dict(issue) -> dict:
        if isinstance(issue, dict):
            return issue
        text = str(issue)
        return {
            "key": text,
            "severity": "info",
            "source": "legacy",
            "confidence": 0,
            "description": text,
            "bbox": [],
            "evidence": [],
            "fix_suggestion": "Legacy issue record; rerun Vision QC v5 to regenerate structured evidence.",
        }

    @classmethod
    def _issues(cls, result: AuditResult) -> list[dict]:
        return [cls._issue_dict(issue) for issue in (result.issues or [])]

    def generate_summary(self) -> dict:
        """生成摘要统计

        Returns:
            包含统计信息的字典
        """
        total = len(self.audit_results)
        pass_count = sum(1 for r in self.audit_results if r.audit_status == "pass")
        fail_count = sum(1 for r in self.audit_results if r.audit_status == "fail")
        need_review_count = sum(1 for r in self.audit_results if r.audit_status == "need_review")
        skipped_count = sum(1 for r in self.audit_results if r.audit_status == "skipped")

        pass_rate = (pass_count / total * 100) if total > 0 else 0.0

        by_status = {
            "pass": pass_count,
            "fail": fail_count,
            "need_review": need_review_count,
            "skipped": skipped_count,
        }

        return {
            "total_audited": total,
            "pass_count": pass_count,
            "fail_count": fail_count,
            "need_review_count": need_review_count,
            "skipped_count": skipped_count,
            "pass_rate": round(pass_rate, 1),
            "by_status": by_status,
        }

    def export_xlsx(self, output_path: Path) -> Path:
        """导出 Excel 报告

        Sheet 1: Summary (总数、通过率)
        Sheet 2: Details (文件路径、基础名称、状态、问题数、critical/major/minor、耗时)
        Sheet 3: Issues (文件路径、issue_key、severity、source、confidence、description)

        Args:
            output_path: 输出文件路径

        Returns:
            输出文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = self.generate_summary()
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        ws_summary["A1"] = "视觉审计报告摘要"
        ws_summary["A1"].font = Font(bold=True, size=14)

        summary_rows = [
            ("审计总数", summary["total_audited"]),
            ("通过", summary["pass_count"]),
            ("失败", summary["fail_count"]),
            ("需复审", summary["need_review_count"]),
            ("已跳过", summary["skipped_count"]),
            ("通过率 (%)", summary["pass_rate"]),
        ]

        for i, (label, value) in enumerate(summary_rows, 3):
            ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=i, column=2, value=value)

        # ===== Sheet 2: Details =====
        ws_details = wb.create_sheet("Details")
        detail_headers = [
            "文件路径", "基础名称", "审计状态", "QC版本",
            "问题总数", "Critical", "Major", "Minor", "耗时(ms)",
        ]
        for col, h in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(self.audit_results, 2):
            issues = self._issues(r)
            issue_count = len(issues)
            critical = sum(1 for i in issues if i.get("severity") == "critical")
            major = sum(1 for i in issues if i.get("severity") == "major")
            minor = sum(1 for i in issues if i.get("severity") == "minor")

            ws_details.cell(row=row_idx, column=1, value=r.file_path)
            ws_details.cell(row=row_idx, column=2, value=r.base_name)
            ws_details.cell(row=row_idx, column=3, value=r.audit_status)
            ws_details.cell(row=row_idx, column=4, value=r.vision_qc_version)
            ws_details.cell(row=row_idx, column=5, value=issue_count)
            ws_details.cell(row=row_idx, column=6, value=critical)
            ws_details.cell(row=row_idx, column=7, value=major)
            ws_details.cell(row=row_idx, column=8, value=minor)
            ws_details.cell(row=row_idx, column=9, value=r.duration_ms)

        # ===== Sheet 3: Issues =====
        ws_issues = wb.create_sheet("Issues")
        issue_headers = [
            "文件路径", "Issue Key", "Severity", "Source", "Confidence", "Description",
        ]
        for col, h in enumerate(issue_headers, 1):
            cell = ws_issues.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        issue_row = 2
        for r in self.audit_results:
            for issue in self._issues(r):
                ws_issues.cell(row=issue_row, column=1, value=r.file_path)
                ws_issues.cell(row=issue_row, column=2, value=issue.get("key", ""))
                ws_issues.cell(row=issue_row, column=3, value=issue.get("severity", ""))
                ws_issues.cell(row=issue_row, column=4, value=issue.get("source", ""))
                ws_issues.cell(row=issue_row, column=5, value=issue.get("confidence", 0))
                ws_issues.cell(row=issue_row, column=6, value=issue.get("description", ""))
                issue_row += 1

        # 自动列宽
        for ws in [ws_summary, ws_details, ws_issues]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value else ""
                        max_len = max(max_len, len(val))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(str(output_path))
        return output_path

    def export_json(self, output_path: Path) -> Path:
        """导出 JSON 报告

        Args:
            output_path: 输出文件路径

        Returns:
            输出文件路径
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        data = {
            "summary": self.generate_summary(),
            "results": [
                {
                    "file_path": r.file_path,
                    "base_name": r.base_name,
                    "audit_status": r.audit_status,
                    "vision_qc_version": r.vision_qc_version,
                    "issues": self._issues(r),
                    "duration_ms": r.duration_ms,
                    "timestamp": r.timestamp,
                }
                for r in self.audit_results
            ],
        }

        output_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return output_path

    def generate_text_report(self) -> str:
        """生成人类可读的文本报告

        Returns:
            文本报告字符串
        """
        summary = self.generate_summary()
        lines = [
            "=" * 60,
            "视觉审计报告",
            "=" * 60,
            "",
            f"审计总数: {summary['total_audited']}",
            f"通过:      {summary['pass_count']}",
            f"失败:      {summary['fail_count']}",
            f"需复审:    {summary['need_review_count']}",
            f"已跳过:    {summary['skipped_count']}",
            f"通过率:    {summary['pass_rate']}%",
            "",
        ]

        # 失败项清单
        failed = [r for r in self.audit_results if r.audit_status == "fail"]
        if failed:
            lines.append("-" * 40)
            lines.append("失败清单:")
            lines.append("-" * 40)
            for r in failed:
                issues = self._issues(r)
                critical = sum(1 for i in issues if i.get("severity") == "critical")
                lines.append(f"  {r.base_name}: {len(issues)} 个问题 (critical={critical})")
                for issue in issues[:5]:
                    lines.append(f"    - [{issue.get('severity', '?')}] {issue.get('key', '')}: {issue.get('description', '')}")
                if len(issues) > 5:
                    lines.append(f"    ... 还有 {len(issues) - 5} 个问题")
            lines.append("")

        # 需复审清单
        need_review = [r for r in self.audit_results if r.audit_status == "need_review"]
        if need_review:
            lines.append("-" * 40)
            lines.append("需复审清单:")
            lines.append("-" * 40)
            for r in need_review:
                issues = self._issues(r)
                lines.append(f"  {r.base_name}: {len(issues)} 个问题")
                for issue in issues[:3]:
                    lines.append(f"    - [{issue.get('severity', '?')}] {issue.get('key', '')}: {issue.get('description', '')}")
            lines.append("")

        lines.append("=" * 60)
        return "\n".join(lines)
